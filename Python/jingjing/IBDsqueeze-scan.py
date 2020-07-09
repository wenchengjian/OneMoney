import numpy as np
import pandas as pd
import talib as ta
import Sendmail
import tushare as ts
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.dates as mdate
import xlrd
from openpyxl import load_workbook

workbook = xlrd.open_workbook('/Users/zhangliwei/myown/US8/IBDresult.xlsx')
sheetnames=workbook.sheet_names()
writer = pd.ExcelWriter('/Users/zhangliwei/myown/US8/scanresult.xlsx',engine='openpyxl')
# sheetnames=['TTD','KL']
signalsheet=None

todaylist=None

for sheet_name in sheetnames:
    # 读取数据
    stock_index = pd.read_excel('/Users/zhangliwei/myown/US8/IBDresult.xlsx', sheet_name=sheet_name, index=False,
                                encoding='utf8')

    # 计算bolling线
    stock_index['upper_b'], stock_index['middle_b'], stock_index['lower_b'] = ta.BBANDS(
        np.asarray(stock_index['Close']),
        timeperiod=20, nbdevup=1.5, nbdevdn=1.5, matype=0)

    ##Keltner Channels
    stock_index['SMA'] = ta.SMA(np.asarray(stock_index['Close']), timeperiod=20)
    stock_index['ATR'] = ta.ATR(np.asarray(stock_index['High']), np.asarray(stock_index['Low']),
                                np.asarray(stock_index['Close']), timeperiod=20)

    stock_index['k_upper'] = stock_index['SMA'] + 1.5 * stock_index['ATR']
    stock_index['k_lower'] = stock_index['SMA'] - 1.5 * stock_index['ATR']

    ##CCI
    stock_index['CCI'] = ta.CCI(np.asarray(stock_index['High']), np.asarray(stock_index['Low']),
                                np.asarray(stock_index['Close']), timeperiod=14)
    # CMO
    stock_index['CMO'] = ta.CMO(np.asarray(stock_index['Close']), timeperiod=14)

    stock_index['sqzIn'] = np.where(
        (stock_index['lower_b'] > stock_index['k_lower']) & (stock_index['upper_b'] < stock_index['k_upper']), 1, 0)
    stock_index['sqzOff'] = np.where(
        (stock_index['lower_b'] < stock_index['k_lower']) | (stock_index['upper_b'] > stock_index['k_upper']), 1, 0)

    # 挤牌信号发出
    stock = stock_index[(stock_index.sqzIn.shift(1) > 0) & (stock_index.sqzOff > 0)]
    # signal
    stock_index['signal'] = np.where(
        (stock_index.sqzIn.shift(1) > 0) & (stock_index.sqzOff > 0) & (stock_index['CCI'] > 0),
        1, 0)
    stock_index['signal'] = np.where(
        (stock_index.sqzIn.shift(1) > 0) & (stock_index.sqzOff > 0) & (stock_index['CCI'] < 0),
        -1, stock_index['signal'])

    # 使用position标记持仓情况，全新的循环法思路；
    position = 0
    # 对每个交易日进行循环
    for item in stock_index.iterrows():  # 逐行遍历；返回的这个item其实一个元组，（label，series）
        # 判断交易信号
        if item[1]['signal'] == 1:
            # 交易信号为1，则记录仓位为1，持有多仓；
            position = 1
            if item[1]['CMO'] < 0:
                position = 0
        elif item[1]['signal'] == -1:
            # 交易信号为-1， 则记录仓位为-1，持有空仓；
            position = -1
            if item[1]['CMO'] > 0:
                position = 0
        else:
            pass  # 啥都不做；
        # 记录每日持仓情况
        stock_index.loc[item[0], 'position'] = position  # 自动往下填充的就是上一个产生的交易信号；关键；
    # 计算股票每日收益率
    stock_index['pct_change'] = stock_index['Close'].pct_change()
    # 计算股票的累积收益率
    stock_index['return'] = (stock_index['pct_change'] + 1).cumprod()
    # 计算策略每日收益率
    stock_index['strategy_return'] = stock_index['position'] * stock_index['pct_change']
    # 计算策略累积收益率
    stock_index['cum_strategy_return'] = (stock_index['strategy_return'] + 1).cumprod()

    signalsheeti = stock_index[['Date','signal','position','return','cum_strategy_return']]
    signalsheeti.insert(0, 'symbol', sheet_name)
    signalsheeti.to_excel(excel_writer=writer, sheet_name=sheet_name,encoding="utf-8", index=True)

    if signalsheeti.iloc[-1:].signal.values==1 :
        addlist=pd.DataFrame(signalsheeti.iloc[-1:])
        if todaylist is None :
            todaylist= addlist
        else :
            todaylist=todaylist.append(addlist,ignore_index=True)


    print('Save ' + sheet_name + ' Complete!')

writer.save()
writer.close()

Sendmail.send_email(u"ahjing99@163.com",u"84635858@qq.com",u"今日挤牌",u"xxx",todaylist.to_html())
Sendmail.send_email(u"ahjing99@163.com",u"cwen@vmware.com",u"今日挤牌",u"xxx",todaylist.to_html())

