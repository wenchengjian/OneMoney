import numpy as np
import pandas as pd
import talib as ta
import Sendmail
import tushare as ts
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.dates as mdate
import xlrd
from openpyxl import load_workbook

workbook = xlrd.open_workbook('/Users/zhangliwei/myown/US8/IBDresult.xlsx')
sheetnames=workbook.sheet_names()
writer = pd.ExcelWriter('/Users/zhangliwei/myown/US8/DailyIBDsqueezescanresult.xlsx',engine='openpyxl')
writer1 = pd.ExcelWriter('/Users/zhangliwei/myown/US8/TODAYLIST-IBD50SQUEEZE.xlsx',engine='openpyxl')
# sheetnames=['KL']
signalsheet=None
returnlist=None
todaylist=pd.DataFrame()

for sheet_name in sheetnames:
    # 读取数据
    stock_index = pd.read_excel('/Users/zhangliwei/myown/US8/IBDresult.xlsx', sheet_name=sheet_name, index=False,
                                encoding='utf8')

    # 计算bolling线
    stock_index['upper_b'], stock_index['middle_b'], stock_index['lower_b'] = ta.BBANDS(
        np.asarray(stock_index['Close']),
        timeperiod=20, nbdevup=2, nbdevdn=2, matype=0)

    ##Keltner Channels
    stock_index['SMA'] = ta.SMA(np.asarray(stock_index['Close']), timeperiod=20)
    stock_index['ATR'] = ta.ATR(np.asarray(stock_index['High']), np.asarray(stock_index['Low']),
                                np.asarray(stock_index['Close']), timeperiod=20)

    stock_index['k_upper'] = stock_index['SMA'] + 1.5 * stock_index['ATR']
    stock_index['k_lower'] = stock_index['SMA'] - 1.5 * stock_index['ATR']

    ##CCI,判断方向
    stock_index['CCI'] = ta.CCI(np.asarray(stock_index['High']), np.asarray(stock_index['Low']),
                                np.asarray(stock_index['Close']), timeperiod=14)
    # CMO，判断动量
    stock_index['CMO'] = ta.CMO(np.asarray(stock_index['Close']), timeperiod=5)
    # stock_index['ADX']=ta.ADX(np.asarray(stock_index['High']), np.asarray(stock_index['Low']),
    #                             np.asarray(stock_index['Close']), timeperiod=14)

    #布林带在肯特纳内
    stock_index['sqzIn'] = np.where(
        (stock_index['lower_b'] > stock_index['k_lower']) & (stock_index['upper_b'] < stock_index['k_upper']), 1, 0)
    # 布林带在肯特纳外
    stock_index['sqzOff'] = np.where(
        (stock_index['lower_b'] < stock_index['k_lower']) | (stock_index['upper_b'] > stock_index['k_upper']), 1, 0)

    # # 挤牌信号发出
    # stock = stock_index[(stock_index.sqzIn.shift(1) > 0) & (stock_index.sqzOff > 0)]
    # 挤牌信号发出
    stock_index['signal'] = np.where(
        (stock_index.sqzIn.shift(1) > 0) & (stock_index.sqzOff > 0) & (stock_index['CCI'] > 0),
        1, 0)
    stock_index['signal'] = np.where(
        (stock_index.sqzIn.shift(1) > 0) & (stock_index.sqzOff > 0) & (stock_index['CCI'] < 0),
        -1, stock_index['signal'])


    # 退出策略
    stock_index['exit'] = np.where(
        (stock_index['CMO'].shift(-1) > 0) & (stock_index['CMO'] < 0) , -1, 0)
    stock_index['exit'] = np.where(
        (stock_index['CMO'].shift(-1) < 0) & (stock_index['CMO'] > 0), 1, stock_index['exit'])

    # 使用position标记持仓情况，全新的循环法思路；
    position = 0
    # 对每个交易日进行循环
    for item in stock_index.iterrows():  # 逐行遍历；返回的这个item其实一个元组，（label，series）
        # 判断交易信号
        if item[1]['signal'] == 1:
            # 交易信号为1，则记录仓位为1，持有多仓；
            position = 1
        elif item[1]['signal'] == -1:
            # 交易信号为-1， 则记录仓位为-1，持有空仓；
            position = -1
        else  :
            if item[1]['exit'] * position < 0:
                position = 0
            else:
                pass
        # 记录每日持仓情况
        stock_index.loc[item[0], 'position'] = position  # 自动往下填充的就是上一个产生的交易信号；关键；



    # # 对每个交易日进行循环退出
    # for iteme in stock_index.iterrows():  # 逐行遍历；返回的这个item其实一个元组，（label，series）
    #     # 判断退出信号
    #     if iteme[1]['exit'] * iteme[1]['position']<0:
    #         # 仓位信号为1，退出信号为-1，平仓；仓位信号为-1，退出信号为1，平仓；
    #         position = 0
    #     else:
    #         pass  # 啥都不做；

    # 计算股票每日收益率
    stock_index['pct_change'] = stock_index['Close'].pct_change()
    # 计算股票的累积收益率
    stock_index['return'] = (stock_index['pct_change'] + 1).cumprod()
    # 计算策略每日收益率
    stock_index['strategy_return'] = stock_index['position'] * stock_index['pct_change']
    # 计算策略累积收益率
    stock_index['cum_strategy_return'] = (stock_index['strategy_return'] + 1).cumprod()

    signalsheeti = stock_index[['Date','Close','signal','position','exit','return','cum_strategy_return']]
    signalsheeti.insert(0, 'symbol', sheet_name)
    signalsheeti.to_excel(excel_writer=writer, sheet_name=sheet_name,encoding="utf-8", index=True)

    if signalsheeti.iloc[-1:].signal.values==1 :
        addlist=pd.DataFrame(signalsheeti.iloc[-1:])
        if todaylist is None :
            todaylist= addlist
        else :
            todaylist=todaylist.append(addlist,ignore_index=True)

    if signalsheeti.iloc[-1:].exit.values!=0 :
        exitlist=pd.DataFrame(signalsheeti.iloc[-1:])
        if todaylist is None :
            todaylist= exitlist
        else :
            todaylist=todaylist.append(exitlist,ignore_index=True)

    if returnlist is None:
        returnlist=signalsheeti.iloc[-1:]
    else:
        returnlist=returnlist.append(signalsheeti.iloc[-1:],ignore_index=True)
    print('Save ' + sheet_name + ' Complete!')

writer.save()
writer.close()
print(todaylist)
todaylist.to_excel(excel_writer=writer1, sheet_name='IBD50SQUEEZE',encoding="utf-8", index=True)
writer1.save()
writer1.close()